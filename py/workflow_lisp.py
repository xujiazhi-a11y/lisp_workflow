"""
Lisp Workflow 解释器
基于 Lisp.py，扩展支持 AI 工作流

新增特性:
- 英文关键字 (define, lambda, if, begin, quote, set!, pipe, map, reduce, filter)
- 大模型调用 (call-llm)
- JSON 处理 (parse-json, to-json, extract-json)
- 管道操作 (pipe, ->)
- 高阶函数 (map, reduce, filter)
"""

import re
import sys
import math
import os
import io
import functools
from typing import Any, Callable, List, Dict, Optional

# 基准路径（脚本所在目录）
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_loaded_files = set()

# ============================================================
# 导入工作流模块
# ============================================================
try:
    from workflow.llm import call_llm, call_llm_with_system
    from workflow.json_utils import parse_json, to_json, extract_json
    from workflow.text_utils import (
        remove_think_tags, format_string, 
        regex_match, regex_replace,
        split_text, join_text, strip_text
    )
    from workflow.ai_services import (
        lisp_ai_text, lisp_ai_image, lisp_ai_video, lisp_video_concat,
    )
    from workflow.user_errors import format_user_error
    WORKFLOW_MODULES_LOADED = True
except ImportError:
    WORKFLOW_MODULES_LOADED = False
    format_user_error = lambda e: str(e)
    print("警告：未找到 workflow 模块，部分功能将不可用")

# ============================================================
# 基础类型定义
# ============================================================

class Symbol(str):
    """符号类型"""
    pass

Number = (int, float)
List = list
String = str

# 符号表（用于符号去重）
SYMBOL_TABLE = {}

def to_symbol(s: str, symbol_table: dict = None) -> Symbol:
    """将字符串转为符号"""
    if symbol_table is None:
        symbol_table = SYMBOL_TABLE
    if s not in symbol_table:
        symbol_table[s] = Symbol(s)
    return symbol_table[s]

# ============================================================
# 保留字定义（支持中英文）
# ============================================================

# 英文保留字
KW_DEFINE = to_symbol('define')
KW_LAMBDA = to_symbol('lambda')
KW_IF = to_symbol('if')
KW_COND = to_symbol('cond')
KW_BEGIN = to_symbol('begin')
KW_QUOTE = to_symbol('quote')
KW_SET = to_symbol('set!')
KW_DEFMACRO = to_symbol('defmacro')
KW_MACROEXPAND = to_symbol('macroexpand')

# 中文保留字（保持兼容）
KW_定义 = to_symbol('定义')
KW_道 = to_symbol('道')
KW_规定 = to_symbol('规定')
KW_如果 = to_symbol('如果')
KW_情况符合 = to_symbol('情况符合')
KW_否则 = to_symbol('否则')
KW_其它情况 = to_symbol('其它情况')
KW_开始 = to_symbol('开始')
KW_引 = to_symbol('引')
KW_赋 = to_symbol('！赋')
KW_定义宏 = to_symbol('定义宏')
KW_宏展开 = to_symbol('宏展开')

# 工作流保留字
KW_PIPE = to_symbol('pipe')
KW_THREADED = to_symbol('->')
KW_MAP = to_symbol('map')
KW_REDUCE = to_symbol('reduce')
KW_FILTER = to_symbol('filter')
KW_LET = to_symbol('let')
KW_EACH = to_symbol('each')

# 中文工作流保留字
KW_批处理 = to_symbol('批处理')
KW_映射 = to_symbol('映射')
KW_归约 = to_symbol('归约')
KW_筛选 = to_symbol('筛选')
KW_过滤 = to_symbol('过滤')
KW_顺序执行 = to_symbol('顺序执行')
KW_令 = to_symbol('令')
KW_遍历 = to_symbol('遍历')
KW_命 = to_symbol('命')
KW_LOAD = to_symbol('load')
KW_引入 = to_symbol('引入')
KW_OR = to_symbol('or')
KW_AND = to_symbol('and')
KW_或 = to_symbol('或')
KW_与 = to_symbol('与')

# 程序终止符
EOF = Symbol('#<eof-object>')

# ============================================================
# 词法分析器（Tokenizer）
# ============================================================

class Tokenizer:
    """词法分析器"""

    def __init__(self, file):
        self.file = file
        self.line = ''

    def next_token(self):
        """读取下一个token"""
        while True:
            # 如果当前行有内容，先处理完
            if self.line == '':
                self.line = self.file.readline()
                if self.line == '':
                    return EOF

            # 替换阶段：仅处理中文引号和分号，【】不在这里替换
            self.line = self.line.replace('”', '”')
            self.line = self.line.replace('；', ';')

            # 跳过空行
            if not self.line.strip():
                self.line = ''
                continue

            # === 首先检查特殊字符 ===
            first_char = self.line[0]
            if first_char in "()'`[]," or first_char in '（）':
                self.line = self.line[1:]
                if first_char == '（':
                    return '('
                elif first_char == '）':
                    return ')'
                return first_char

            # 【 和 】作为括号直接处理（不在行级替换，避免破坏字符串中的内容）
            if first_char == '【':
                self.line = self.line[1:]
                return '('
            if first_char == '】':
                self.line = self.line[1:]
                return ')'

            # === 跳过开头空白 ===
            if first_char.isspace():
                self.line = self.line[1:]
                continue

            # === 跳过注释 ===
            if first_char == ';':
                self.line = ''
                continue

            # === 处理字符串 ===
            if first_char == '"':
                rest = self.line[1:]  # Everything after the first quote

                # 查找闭合引号（只检查转义和引号，不因特殊字符中断）
                has_closing = False
                closing_pos = -1
                i = 0
                while i < len(rest):
                    ch = rest[i]
                    if ch == '\\':
                        i += 2  # 跳过转义字符
                        continue
                    if ch == '"':
                        has_closing = True
                        closing_pos = i + 1
                        break
                    i += 1

                if has_closing:
                    # Token includes: opening quote (1 char) + content + closing quote
                    # closing_pos is in rest, so total token length = 1 + closing_pos
                    token = self.line[:1 + closing_pos]
                    self.line = self.line[1 + closing_pos:]
                    return token

                # 当前行没有闭合引号，使用多行字符串处理器
                in_string = True
                string_parts = []
                current_line = rest

                while in_string:
                    j = 0
                    while j < len(current_line):
                        ch = current_line[j]
                        if ch == '\\':
                            j += 2
                            continue
                        if ch == '"':
                            # 找到闭合引号
                            string_parts.append(current_line[:j+1])
                            self.line = current_line[j+1:]
                            in_string = False
                            break
                        j += 1

                    if in_string:
                        string_parts.append(current_line)
                        string_parts.append('\n')
                        self.line = ''

                        next_line = self.file.readline()
                        if next_line == '':
                            raise SyntaxError('字符串未闭合')
                        current_line = next_line

                return '"' + ''.join(string_parts)

            # === 解析标识符 ===
            if first_char.isalpha():
                i = 1
                while i < len(self.line) and not self.line[i].isspace() and self.line[i] not in "()'`,\";{}[]" and self.line[i] not in '（）【】':
                    i += 1
                token = self.line[:i]
                self.line = self.line[i:]
                return token

            # === 解析其他标识符 ===
            i = 1
            while i < len(self.line) and not self.line[i].isspace() and self.line[i] not in "()'`,\";{}[]" and self.line[i] not in '（）【】':
                i += 1
            token = self.line[:i]
            self.line = self.line[i:]
            return token

            # === 查找字符串 ===
            in_string = False
            string_start = -1
            i = 0
            while i < len(self.line):
                ch = self.line[i]
                if ch == '\\':
                    i += 2
                    continue
                if ch == '"':
                    if not in_string:
                        string_start = i
                        in_string = True
                    else:
                        token = self.line[string_start:i+1]
                        self.line = self.line[i+1:]
                        return token
                i += 1

            # 如果在字符串中但字符串未闭合
            if in_string:
                raise SyntaxError('字符串未闭合: ' + self.line)

            # 没有找到字符串，也没有特殊字符
            # === 解析标识符或运算符 ===
            # 如果首字符不是字母数字，需要单独处理
            if not first_char.isalnum() and first_char not in "_-":
                # 这是一个运算符或特殊字符
                self.line = self.line[1:]
                return first_char

            # === 解析标识符 ===
            while i < len(self.line) and not self.line[i].isspace() and self.line[i] not in "()'`,\";{}[]":
                i += 1
            token = self.line[:i]
            self.line = self.line[i:]
            return token

# ============================================================
# 语法解析器（Parser）
# ============================================================

def parse(tokens: Tokenizer):
    """解析token流为AST"""

    def read(token):
        if token == '(' or token == '[':
            close = ']' if token == '[' else ')'
            lst = []
            while True:
                token = tokens.next_token()
                if token == close:
                    return lst
                elif str(token) == '#<eof-object>':
                    raise SyntaxError('程序异常终止')
                else:
                    lst.append(read(token))
        elif token == ')' or token == ']':
            raise SyntaxError('多余的 )')
        elif str(token) == '#<eof-object>':
            raise SyntaxError('程序异常终止')
        else:
            return atom(token)

    next_token = tokens.next_token()
    if str(next_token) == '#<eof-object>':
        return EOF
    if next_token == '(' or next_token == '[':
        close = ']' if next_token == '[' else ')'
        lst = []
        while True:
            token = tokens.next_token()
            if token == close:
                return lst
            elif str(token) == '#<eof-object>':
                raise SyntaxError('程序异常终止')
            else:
                lst.append(read(token))
    else:
        return atom(next_token)


def atom(token: str):
    """将token转换为原子值"""
    if token == '#t' or token == '#true' or token == '#真':
        return True
    elif token == '#f' or token == '#false' or token == '#假':
        return False
    elif token.startswith('"'):
        s = token[1:-1]
        # 支持常见的转义序列
        s = s.replace('\\n', '\n').replace('\\t', '\t').replace('\\r', '\r')
        s = s.replace('\\\\', '\\')
        s = s.replace('\\"', '"')
        return s

    # 尝试解析为数字
    try:
        return int(token)
    except ValueError:
        try:
            return float(token)
        except ValueError:
            # 复数: 仅当 token 以数字开头或形如 "数字+i" 时才尝试
            if len(token) > 1 and 'i' in token and token[0].isdigit():
                try:
                    return complex(token.replace('i', 'j', 1))
                except ValueError:
                    pass
            return to_symbol(token)

# ============================================================
# 环境与求值
# ============================================================

class Env(dict):
    """环境：存储变量绑定"""
    
    def __init__(self, params=(), args=(), outer=None):
        self.outer = outer
        
        # 如果参数是单个符号，接收所有实参作为列表
        if isinstance(params, Symbol):
            self.update({params: list(args)})
        else:
            if len(params) != len(args):
                raise TypeError(
                    f'形参个数({len(params)})和实参个数({len(args)})不匹配'
                )
            self.update(zip(params, args))
    
    def find(self, var):
        """查找变量所在的环境层"""
        if var in self:
            return self
        elif self.outer is None:
            raise LookupError(f'未定义的符号: {var}')
        else:
            return self.outer.find(var)

    def set_local(self, var, value):
        """设置变量，如果存在则更新，否则创建新的绑定"""
        if var in self:
            self[var] = value
        else:
            self[var] = value


class Procedure:
    """用户定义的过程（lambda）"""

    def __init__(self, params, body, env):
        self.params = params
        # 确保 body 中的符号是 Symbol 类型
        # 但保留字符串类型（用于字符串字面量）
        self.body = self._normalize_body(body)
        self.env = env

    def _normalize_body(self, body):
        """将 body 中的字符串（标识符）转换为 Symbol，保留字符串（字面量）"""
        def convert(item):
            if isinstance(item, str):
                # 保留字符串类型，不转换为 Symbol
                # 这样字符串字面量会被当作常量返回
                return item
            elif isinstance(item, list):
                return [convert(x) for x in item]
            return item
        return convert(body) if isinstance(body, list) else body

    def __call__(self, *args):
        return evaluate(self.body, Env(self.params, args, self.env))


# ============================================================
# 内置函数定义
# ============================================================

def _template_format(template, *args):
    if '{{' in template:
        result = template
        for i in range(0, len(args) - 1, 2):
            result = result.replace('{{' + str(args[i]) + '}}', str(args[i + 1]))
        return result
    return template % args


def _list_extend(first, rest):
    """原地扩展列表，避免 O(n) 拷贝。与 put 一致采用可变语义。"""
    if not isinstance(first, list):
        first = [first]
    for lst in rest:
        if isinstance(lst, list):
            first.extend(lst)
        else:
            first.append(lst)
    return first


def _safe_car(lst):
    if not isinstance(lst, list) or len(lst) == 0:
        raise IndexError('列表为空，无法取首元素')
    return lst[0]


def _safe_cdr(lst):
    if not isinstance(lst, list):
        raise TypeError('cdr 需要列表参数')
    if len(lst) == 0:
        raise IndexError('列表为空，无法取剩余元素')
    return lst[1:]


def _safe_read_file(path):
    full = os.path.join(BASE_DIR, path)
    try:
        with open(full, 'r', encoding='utf-8') as f:
            return f.read()
    except FileNotFoundError:
        raise FileNotFoundError(f'文件不存在：{path}') from None
    except OSError:
        raise OSError(f'读取文件失败：{path}') from None


def _safe_write_file(path, content):
    full = os.path.join(BASE_DIR, path)
    try:
        parent = os.path.dirname(full)
        if parent:
            os.makedirs(parent, exist_ok=True)
        with open(full, 'w', encoding='utf-8') as f:
            f.write(content)
    except OSError:
        raise OSError(f'写入文件失败：{path}') from None


def make_global_env():
    """创建全局环境"""
    env = Env()
    
    # 数学函数
    env.update(vars(math))
    
    # 基本运算
    env.update({
        '+': lambda *args: sum(args),
        '-': lambda *args: args[0] - sum(args[1:]) if len(args) > 1 else -args[0],
        '*': lambda *args: functools.reduce(lambda a, b: a * b, args, 1),
        '/': lambda *args: functools.reduce(lambda a, b: a / b, args) if len(args) > 1 else 1/args[0],
        'mod': lambda a, b: a % b,
        '%': lambda a, b: a % b,
        
        # 比较运算
        '>': lambda a, b: a > b,
        '<': lambda a, b: a < b,
        '>=': lambda a, b: a >= b,
        '<=': lambda a, b: a <= b,
        '=': lambda a, b: a == b,
        'eq?': lambda a, b: a is b,
        'equal?': lambda a, b: a == b,
        
        # 布尔运算
        'and': lambda *args: all(args),
        'or': lambda *args: any(args),
        'not': lambda x: not x,
        
        # 列表操作
        'list': lambda *args: list(args),
        'cons': lambda a, b: [a] + (b if isinstance(b, list) else [b]),
        'car': _safe_car,
        'cdr': _safe_cdr,
        'first': _safe_car,
        'rest': _safe_cdr,
        'length': len,
        'append': lambda first, *rest: _list_extend(first, rest),
        'reverse': lambda lst: lst[::-1],
        'nth': lambda n, lst: lst[n],
        'take': lambda n, lst: lst[:n],
        'drop': lambda n, lst: lst[n:],
        'sort': lambda fn, lst: sorted(lst, key=functools.cmp_to_key(fn)),
        'member?': lambda item, lst: item in lst,
        'to-number': lambda x: int(x) if isinstance(x, int) else (float(x) if isinstance(x, (float, str)) else 0),

        # 数学函数（按博客设计）
        'square': lambda x: x * x,
        'sqrt': lambda x: x ** 0.5,
        'expt': lambda base, exp: base ** exp,
        'abs': abs,
        'min': min,
        'max': max,
        'average': lambda *args: sum(args) / len(args) if args else 0,
        'even?': lambda x: x % 2 == 0,
        'odd?': lambda x: x % 2 != 0,
        'prime?': lambda n: n > 1 and all(n % i != 0 for i in range(2, int(n**0.5) + 1)),
        'divides?': lambda a, b: b % a == 0,

        # 类型判断
        'null?': lambda x: x == [] or x is None,
        'list?': lambda x: isinstance(x, list),
        'number?': lambda x: isinstance(x, Number),
        'string?': lambda x: isinstance(x, str) and not isinstance(x, Symbol),
        'symbol?': lambda x: isinstance(x, Symbol),
        'boolean?': lambda x: isinstance(x, bool),
        'procedure?': callable,
        'empty?': lambda x: len(x) == 0 if hasattr(x, '__len__') else False,
        'dict': lambda *args: {args[i]: args[i+1] for i in range(0, len(args), 2)},
        'dict?': lambda x: isinstance(x, dict),
        
        # 字典操作
        'get': lambda d, k: d.get(k) if isinstance(d, dict) else None,
        'put': lambda d, k, v: (d.update({k: v}), d)[1] if isinstance(d, dict) else d,
        'keys': lambda d: list(d.keys()) if isinstance(d, dict) else [],
        'values': lambda d: list(d.values()) if isinstance(d, dict) else [],
        
        # 字符串操作
        'str': lambda *args: ''.join(str(a) for a in args),
        'str-concat': lambda *args: ''.join(str(a) for a in args),
        'str-join': lambda sep, lst: sep.join(str(x) for x in lst),
        'str-split': lambda sep, s: s.split(sep),
        'str-replace': lambda old, new, s: s.replace(old, new),
        'str-trim': lambda s: s.strip(),
        'str-upper': lambda s: s.upper(),
        'str-lower': lambda s: s.lower(),
        'str-starts?': lambda prefix, s: s.startswith(prefix),
        'str-ends?': lambda suffix, s: s.endswith(suffix),
        'str-contains?': lambda substr, s: substr in s,
        'format': lambda template, *args: _template_format(template, *args),
        
        # 打印输出
        'print': lambda *args: print(' '.join(to_lisp_str(x) if not isinstance(x, str) else x for x in args)),
        'println': lambda *args: print(' '.join(to_lisp_str(x) if not isinstance(x, str) else x for x in args)),
        'pr': lambda x: print(x, end=''),
        
        # I/O - 使用固定基准路径
        'read-file': _safe_read_file,
        'write-file': _safe_write_file,
    })
    
    # ============================================================
    # 高阶函数
    # ============================================================
    
    def lisp_map(fn, lst):
        """map: 对列表每项应用函数"""
        return [fn(x) for x in lst]
    
    def lisp_reduce(fn, init, lst):
        """reduce: 折叠列表"""
        return functools.reduce(fn, lst, init)
    
    def lisp_filter(fn, lst):
        """filter: 过滤列表"""
        return [x for x in lst if fn(x)]
    
    def lisp_pipe(init, *fns):
        """pipe: 管道操作，值依次流经各函数"""
        result = init
        for fn in fns:
            result = fn(result)
        return result

    def lisp_threaded(init, *forms):
        """->: 线程宏，将值作为第一个参数传入后续每个表达式"""
        result = init
        for form in forms:
            if isinstance(form, list):
                if len(form) > 0:
                    # (f args...) -> (f result args...)
                    fn = evaluate(form[0], env)
                    args = [evaluate(a, env) for a in form[1:]]
                    result = fn(result, *args)
            else:
                fn = evaluate(form, env)
                result = fn(result)
        return result
    
    env.update({
        'map': lisp_map,
        'reduce': lisp_reduce,
        'filter': lisp_filter,
        'pipe': lisp_pipe,
        '->': lisp_threaded,
        'each': lambda fn, lst: [fn(x) for x in lst][-1],  # 执行副作用
    })
    
    # ============================================================
    # JSON 处理
    # ============================================================
    
    if WORKFLOW_MODULES_LOADED:
        env.update({
            'parse-json': parse_json,
            'to-json': lambda data: to_json(data),
            'extract-json': lambda s, d=None: extract_json(s, default=d),
        })
    else:
        import json
        env.update({
            'parse-json': json.loads,
            'to-json': lambda data: json.dumps(data, ensure_ascii=False, indent=2),
            'extract-json': lambda s: json.loads(re.search(r'(\[.*\]|\{.*\})', s, re.DOTALL).group(1)),
        })
    
    # ============================================================
    # 文本处理
    # ============================================================
    
    if WORKFLOW_MODULES_LOADED:
        env.update({
            'remove-think': remove_think_tags,
            'regex-match': regex_match,
            'regex-replace': regex_replace,
            'strip': strip_text,
        })
    else:
        env.update({
            'remove-think': lambda s: re.sub(r'<(think|thinking|thought)[^>]*>.*?</\1\s*>', '', s, flags=re.DOTALL).strip(),
            'regex-match': lambda p, s: re.search(p, s).group(0) if re.search(p, s) else None,
            'regex-replace': lambda p, r, s: re.sub(p, r, s),
            'strip': lambda s: s.strip(),
        })
    
    # ============================================================
    # 大模型调用
    # ============================================================
    
    if WORKFLOW_MODULES_LOADED:
        env.update({
            'call-llm': call_llm,
            'llm': call_llm,  # 别名
            'ai-text': lisp_ai_text,
            'ai-image': lisp_ai_image,
            'ai-video': lisp_ai_video,
            'video-concat': lisp_video_concat,
        })
    else:
        def mock_llm(prompt):
            return f"[模拟 LLM 响应] {prompt[:50]}..."
        def mock_ai_text(provider, model, prompt, *kw):
            return f"[模拟 AI 文本] provider={provider} model={model} prompt={str(prompt)[:50]}..."
        def mock_ai_image(provider, model, prompt, *kw):
            return f"[模拟 AI 图像] provider={provider} model={model} prompt={str(prompt)[:30]}..."
        def mock_ai_video(provider, first_frame, last_frame, prompt, duration):
            return f"[模拟 AI 视频] provider={provider} duration={duration} prompt={str(prompt)[:30]}..."
        def mock_video_concat(video_paths, output_path):
            return f"[模拟视频拼接] output={output_path}"
        env.update({
            'call-llm': mock_llm,
            'llm': mock_llm,
            'ai-text': mock_ai_text,
            'ai-image': mock_ai_image,
            'ai-video': mock_ai_video,
            'video-concat': mock_video_concat,
        })
    
    return env


# 全局环境
GLOBAL_ENV = make_global_env()

# 宏表：存储用户定义的宏
MACRO_TABLE = {}
_CN_MACRO_ALIASES = {}  # 中文宏别名


def _macro_expand(macro_def, args):
    """
    简单宏展开：
    - macro_def = [params, template]
    - args = 实参列表
    - 返回：替换后的模板
    """
    params, template = macro_def
    # 参数绑定
    bindings = {}
    if isinstance(params, list):
        for i, p in enumerate(params):
            if i < len(args):
                bindings[p] = args[i]
    elif isinstance(params, Symbol):
        # 单个参数，接收所有实参
        bindings[params] = args

    # 递归替换模板中的参数
    def substitute(expr):
        if isinstance(expr, Symbol):
            if expr in bindings:
                return bindings[expr]
            return expr
        elif isinstance(expr, list):
            return [substitute(e) for e in expr]
        else:
            return expr

    return substitute(template)

# 飞书 Webhook 配置（使用环境变量或占位符）
import os
FEISHU_WEBHOOK = os.environ.get('FEISHU_WEBHOOK', '[请配置飞书Webhook地址]')

def feishu_send(title, content):
    """发送消息到飞书群"""
    try:
        import urllib.request
        import json
        card = {
            "msg_type": "interactive",
            "card": {
                "config": {"wide_screen_mode": True},
                "header": {
                    "title": {"tag": "plain_text", "content": title},
                    "template": "red" if "负向" in title else "green"
                },
                "elements": [
                    {"tag": "div", "text": {"tag": "plain_text", "content": content}}
                ]
            }
        }
        data = json.dumps(card).encode('utf-8')
        req = urllib.request.Request(FEISHU_WEBHOOK, data=data, headers={'Content-Type': 'application/json'})
        with urllib.request.urlopen(req, timeout=10) as response:
            result = json.loads(response.read().decode('utf-8'))
            return "飞书发送成功" if result.get('code') == 0 or result.get('StatusCode') == 0 else f"发送失败: {result}"
    except Exception as e:
        return f"飞书发送失败: {str(e)}"

GLOBAL_ENV['send-to-feishu'] = feishu_send
GLOBAL_ENV['wait-seconds'] = lambda s: __import__('time').sleep(max(0, float(s)))
GLOBAL_ENV[to_symbol('等待秒数')] = GLOBAL_ENV['wait-seconds']

# 交互式用户输入函数
def _user_input(prompt=""):
    """读取用户输入（用于交互式工作流）"""
    try:
        return input(prompt)
    except (EOFError, KeyboardInterrupt):
        return ""

GLOBAL_ENV['user-input'] = _user_input
GLOBAL_ENV[to_symbol('用户输入')] = _user_input

# ============================================================
# 浏览器自动化原语（Playwright）
# ============================================================

def _browser_start():
    """启动浏览器，返回 (browser, context) 元组"""
    try:
        from playwright.sync_api import sync_playwright
        p = sync_playwright().start()
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        return {'browser': browser, 'context': context, 'playwright': p}
    except ImportError:
        raise RuntimeError("未安装 playwright，请运行: pip install playwright && python -m playwright install chromium")

def _browser_open(browser_ctx, url):
    """在浏览器中打开页面"""
    page = browser_ctx['context'].new_page()
    page.goto(url, wait_until="domcontentloaded", timeout=15000)
    __import__('time').sleep(3)
    return page

def _page_wait_for_login(page, timeout=120, interactive=False):
    """等待用户完成登录

    参数:
        page: Playwright页面对象
        timeout: 超时时间（秒）
        interactive: 是否启用交互模式（用户按回车确认）
    """
    import time

    if interactive:
        # 交互模式：等待用户按回车确认
        try:
            input()
        except (EOFError, KeyboardInterrupt):
            pass
        return True

    # 自动检测模式：等待登录框消失（radio按钮出现后再消失），或等待URL变化
    # 先等待登录框出现（radio为选择单位弹窗）
    start_time = time.time()
    radio_appeared = False
    while (time.time() - start_time) < timeout:
        try:
            count = page.locator('.el-radio__inner').count()
            if count > 0:
                radio_appeared = True
                break
        except Exception:
            pass
        time.sleep(0.5)

    if not radio_appeared:
        # 没检测到选择单位弹窗，回退到URL检测
        login_urls = ['/unitLogin', '/login', '#/unitLogin', '#/login']
        while (time.time() - start_time) < timeout:
            current_url = page.url
            if any(url_part in current_url for url_part in login_urls):
                # 等待离开登录页
                while (time.time() - start_time) < timeout:
                    current_url = page.url
                    if not any(url_part in current_url for url_part in login_urls):
                        time.sleep(1)
                        return True
                    time.sleep(0.5)
            time.sleep(0.5)
        return False

    # 检测到选择单位弹窗，等待它消失（用户完成选择确认）
    while (time.time() - start_time) < timeout:
        try:
            count = page.locator('.el-radio__inner').count()
            if count == 0:
                time.sleep(1)
                return True
        except Exception:
            pass
        time.sleep(0.5)

    return False

def _page_find(page, selector, attrs=None):
    """在页面中查找元素"""
    try:
        if attrs and isinstance(attrs, dict):
            if 'placeholder' in attrs:
                loc = page.locator(f"{selector}[placeholder='{attrs['placeholder']}']")
                if loc.count() > 0:
                    return loc.first
            if 'text' in attrs:
                # 策略1: 直接CSS :has-text
                loc = page.locator(f"{selector}:has-text('{attrs['text']}')")
                if loc.count() > 0:
                    return loc.first
                # 策略2: 任意元素 :has-text
                loc = page.locator(f"*:has-text('{attrs['text']}')")
                if loc.count() > 0:
                    return loc.first
                # 策略3: 遍历所有匹配selector的元素，精确匹配文本
                all_elems = page.locator(selector).all()
                for elem in all_elems:
                    try:
                        text = elem.inner_text()
                        if attrs['text'] in text:
                            return elem
                    except:
                        pass
                # 策略4: 如果selector是通用标签，尝试常见交互元素
                if selector in ('*', 'div', 'span'):
                    for tag in ['button', 'a', 'li', 'div', 'span']:
                        loc = page.locator(f"{tag}:has-text('{attrs['text']}')")
                        if loc.count() > 0:
                            return loc.first
        else:
            loc = page.locator(selector)
            if loc.count() > 0:
                return loc.first
    except Exception as e:
        pass
    return None

def _page_find_all(page, selector):
    """查找所有匹配元素"""
    try:
        return page.locator(selector).all()
    except:
        return []

def _elem_fill(elem, value):
    """填写元素"""
    try:
        elem.fill(str(value))
        return True
    except:
        return False

def _elem_click(elem):
    """点击元素"""
    try:
        elem.click()
        return True
    except:
        return False

def _page_exec(page, script):
    """在页面中执行JS"""
    try:
        return page.evaluate(script)
    except Exception as e:
        return f"JS执行失败: {e}"

def _page_click(page, selector):
    """通过CSS选择器点击元素（用于失焦等操作）"""
    try:
        loc = page.locator(selector)
        if loc.count() > 0:
            loc.first.click()
            return True
        return False
    except:
        return False

def _page_click_text(page, text):
    """点击包含指定文本的元素（用于失焦等操作）"""
    try:
        loc = page.locator(f"text={text}")
        if loc.count() > 0:
            loc.first.click()
            return True
        return False
    except:
        return False

def _page_wait_for_selector(page, selector, timeout=10000):
    """等待元素出现"""
    try:
        page.wait_for_selector(selector, timeout=timeout)
        return True
    except:
        return False

def _page_check(page, selector):
    """勾选复选框"""
    try:
        loc = page.locator(selector)
        if loc.count() > 0:
            loc.first.check()
            return True
        return False
    except:
        return False

def _page_click_checkbox_by_label(page, label):
    """通过标签文本点击对应的复选框"""
    try:
        # 先找 label 元素
        loc = page.locator(f"label:has-text('{label}')")
        if loc.count() > 0:
            loc.first.click()
            return True
        # 再找包含文本的 span/div
        loc = page.locator(f"span:has-text('{label}')")
        if loc.count() > 0:
            parent = loc.first.locator("..")
            checkbox = parent.locator("input[type=checkbox]")
            if checkbox.count() > 0:
                checkbox.first.check()
                return True
            parent.click()
            return True
        return False
    except:
        return False

def _page_screenshot(page, path):
    """截图保存"""
    try:
        full = os.path.join(BASE_DIR, path)
        os.makedirs(os.path.dirname(full), exist_ok=True)
        page.screenshot(path=full)
        return full
    except Exception as e:
        return f"截图失败: {e}"

def _browser_close(browser_ctx):
    """关闭浏览器"""
    try:
        browser_ctx['browser'].close()
        browser_ctx['playwright'].stop()
        return True
    except:
        return False

def _excel_read(path):
    """读取Excel文件，返回 [表头, 数据行列表]"""
    try:
        from openpyxl import load_workbook
        full = os.path.join(BASE_DIR, path)
        wb = load_workbook(full, data_only=True)
        ws = wb.active
        
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers.append(str(header))
        
        rows = []
        for row in range(2, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                row_data.append(val)
            rows.append(row_data)
        
        return [headers, rows]
    except ImportError:
        raise RuntimeError("未安装 openpyxl，请运行: pip install openpyxl")
    except Exception as e:
        raise RuntimeError(f"读取Excel失败: {e}")

def _format_date(date_val):
    """格式化日期"""
    if hasattr(date_val, 'strftime'):
        return date_val.strftime('%Y-%m-%d')
    return str(date_val)

def _page_scan_elements(page, output_path=None):
    """扫描页面中所有可交互元素，返回结构化数据"""
    try:
        # 先滚动页面以加载懒加载内容
        page.evaluate("""
        () => {
            return new Promise((resolve) => {
                let totalHeight = 0;
                const distance = 300;
                const timer = setInterval(() => {
                    const scrollHeight = document.body.scrollHeight;
                    window.scrollBy(0, distance);
                    totalHeight += distance;
                    if (totalHeight >= scrollHeight) {
                        clearInterval(timer);
                        window.scrollTo(0, 0);
                        resolve('scrolled');
                    }
                }, 100);
                // 最多滚动3秒
                setTimeout(() => {
                    clearInterval(timer);
                    window.scrollTo(0, 0);
                    resolve('timeout');
                }, 3000);
            });
        }
        """)
        
        # 等待一下让懒加载内容渲染
        __import__('time').sleep(0.5)
        
        script = """
        () => {
            const interactiveTags = ['A', 'BUTTON', 'INPUT', 'SELECT', 'TEXTAREA', 'LI', 'DIV', 'SPAN'];
            const results = [];
            const seen = new Set();
            
            function getElementInfo(el) {
                const tag = el.tagName;
                const text = (el.innerText || el.textContent || '').trim().substring(0, 100);
                const placeholder = el.getAttribute('placeholder') || '';
                const cls = el.className || '';
                const id = el.id || '';
                const type = el.getAttribute('type') || '';
                const href = el.getAttribute('href') || '';
                const role = el.getAttribute('role') || '';
                
                // 生成推荐选择器
                let selector = tag.toLowerCase();
                if (id) selector += '#' + id;
                else if (cls) {
                    const classes = cls.split(' ').filter(c => c);
                    // 使用更具体的类名
                    const specificClass = classes.find(c => 
                        !c.startsWith('el-icon') && 
                        !c.startsWith('is-') && 
                        c.length > 2
                    ) || classes[0];
                    if (specificClass) selector += '.' + specificClass;
                }
                
                // 判断元素是否可见且可交互
                const rect = el.getBoundingClientRect();
                const style = window.getComputedStyle(el);
                const isVisible = rect.width > 0 && rect.height > 0 && 
                                  style.display !== 'none' && 
                                  style.visibility !== 'hidden';
                
                return {
                    tag: tag,
                    text: text,
                    placeholder: placeholder,
                    class: cls,
                    id: id,
                    type: type,
                    href: href,
                    role: role,
                    selector: selector,
                    isVisible: isVisible,
                    x: Math.round(rect.x),
                    y: Math.round(rect.y),
                    width: Math.round(rect.width),
                    height: Math.round(rect.height)
                };
            }
            
            // 扫描所有可交互标签
            interactiveTags.forEach(tag => {
                document.querySelectorAll(tag).forEach(el => {
                    const info = getElementInfo(el);
                    const key = info.tag + '|' + info.text + '|' + info.selector;
                    if (!seen.has(key) && info.isVisible) {
                        seen.add(key);
                        results.push(info);
                    }
                });
            });
            
            // 额外扫描带点击事件的元素
            document.querySelectorAll('[onclick], [role="button"]').forEach(el => {
                const info = getElementInfo(el);
                const key = info.tag + '|' + info.text + '|' + info.selector;
                if (!seen.has(key) && info.isVisible) {
                    seen.add(key);
                    results.push(info);
                }
            });
            
            return {
                url: window.location.href,
                title: document.title,
                timestamp: new Date().toISOString(),
                elementCount: results.length,
                elements: results
            };
        }
        """
        result = page.evaluate(script)
        
        # 如果指定了输出路径，保存到文件
        if output_path:
            import json
            full_path = os.path.join(BASE_DIR, output_path)
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
            with open(full_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=2)
        
        return result
    except Exception as e:
        return {"error": str(e), "url": "", "title": "", "elements": []}

# 注册浏览器自动化函数
GLOBAL_ENV['browser-start'] = _browser_start
GLOBAL_ENV[to_symbol('浏览器启动')] = _browser_start
GLOBAL_ENV['browser-open'] = _browser_open
GLOBAL_ENV[to_symbol('浏览器打开')] = _browser_open
GLOBAL_ENV['page-find'] = _page_find
GLOBAL_ENV[to_symbol('页面查找')] = _page_find
GLOBAL_ENV['page-find-all'] = _page_find_all
GLOBAL_ENV[to_symbol('页面查找所有')] = _page_find_all
GLOBAL_ENV['elem-fill'] = _elem_fill
GLOBAL_ENV[to_symbol('元素填写')] = _elem_fill
GLOBAL_ENV['elem-click'] = _elem_click
GLOBAL_ENV[to_symbol('元素点击')] = _elem_click
GLOBAL_ENV['page-exec'] = _page_exec
GLOBAL_ENV[to_symbol('页面执行')] = _page_exec
GLOBAL_ENV['page-screenshot'] = _page_screenshot
GLOBAL_ENV[to_symbol('页面截图')] = _page_screenshot
GLOBAL_ENV['browser-close'] = _browser_close
GLOBAL_ENV[to_symbol('浏览器关闭')] = _browser_close
GLOBAL_ENV['excel-read'] = _excel_read
GLOBAL_ENV[to_symbol('Excel读取')] = _excel_read
GLOBAL_ENV['format-date'] = _format_date
GLOBAL_ENV[to_symbol('格式化日期')] = _format_date
GLOBAL_ENV['page-scan'] = _page_scan_elements
GLOBAL_ENV[to_symbol('扫描页面元素')] = _page_scan_elements
GLOBAL_ENV['page-wait-login'] = _page_wait_for_login
GLOBAL_ENV[to_symbol('等待登录跳转')] = _page_wait_for_login
GLOBAL_ENV['page-click'] = _page_click
GLOBAL_ENV[to_symbol('点击元素')] = _page_click
GLOBAL_ENV['page-click-text'] = _page_click_text
GLOBAL_ENV[to_symbol('点击文本')] = _page_click_text
GLOBAL_ENV['page-wait-selector'] = _page_wait_for_selector
GLOBAL_ENV[to_symbol('等待元素出现')] = _page_wait_for_selector
GLOBAL_ENV['page-check'] = _page_check
GLOBAL_ENV[to_symbol('勾选')] = _page_check
GLOBAL_ENV['page-click-checkbox'] = _page_click_checkbox_by_label
GLOBAL_ENV[to_symbol('勾选文本')] = _page_click_checkbox_by_label

# ============================================================
# 中文别名（汉化兼容层）
# ============================================================

_CN_ALIASES = {
    # 输出（按博客设计）
    '打印': 'print',
    '输出': 'print',
    '显示': 'print',
    '换行': 'println',
    # 字符串
    '格式化': 'format',
    '文本拼接': 'str-concat',
    '文本连接': 'str-join',
    '文本裁剪': 'str-trim',
    '文本包含': 'str-contains?',
    '文本开头': 'str-starts?',
    '文本结尾': 'str-ends?',
    '文本替换': 'str-replace',
    '文本分割': 'str-split',
    '转文本': 'str',
    # 列表（按博客设计）
    '序列': 'list',
    '长度': 'length',
    '合并': 'append',
    '追加': 'append',
    '反转': 'reverse',
    '前项': 'car',
    '前头的项': 'car',
    '第一项': 'car',
    '后项': 'cdr',
    '后面的项': 'cdr',
    '序对': 'cons',
    '取项': 'nth',
    # 高阶函数（按博客设计）
    '遍历': 'each',
    '批处理': 'map',
    '映射': 'map',
    '归约': 'reduce',
    '顺序执行': 'reduce',
    '筛选': 'filter',
    '过滤': 'filter',
    # 类型判断（按博客设计）
    '为空?': 'null?',
    '到空值了?': 'null?',
    '成对?': 'pair?',
    '是列表?': 'list?',
    '是数字?': 'number?',
    '是文本?': 'string?',
    '是质数?': 'prime?',
    '整除?': 'divides?',
    '偶数?': 'even?',
    # 逻辑
    '与': 'and',
    '或': 'or',
    '非': 'not',
    '不': 'not',
    # 数学（按博客设计）
    '取余': 'mod',
    '取余数': 'mod',
    '加': '+',
    '减': '-',
    '乘': '*',
    '除': '/',
    '平方': 'square',
    '开方': 'sqrt',
    '幂': 'expt',
    '取最小值': 'min',
    '取最大值': 'max',
    '取平均': 'average',
    # I/O
    '读文件': 'read-file',
    '写文件': 'write-file',
    # 字典
    '字典': 'dict',
    '取值': 'get',
    '赋值': 'put',
    '键集': 'keys',
    '值集': 'values',
    # 排序与成员
    '排序': 'sort',
    '包含?': 'member?',
    '转数字': 'to-number',
    '大写': 'str-upper',
    '小写': 'str-lower',
    '取前': 'take',
    # 工作流标准库
    '调用模型': 'call-llm',
    'AI文本': 'ai-text',
    'AI图像': 'ai-image',
    'AI视频': 'ai-video',
    '视频拼接': 'video-concat',
    '静帧拼视频': 'slideshow-video',
    '用户输入': 'user-input',
    '用户输入数据': 'client-input',  # 中文别名，指向执行时注入的 client-input
    '去除思考': 'remove-think',
    '解析JSON': 'parse-json',
    '转JSON': 'to-json',
    '提取JSON': 'extract-json',
}

for cn_name, en_name in _CN_ALIASES.items():
    if en_name in GLOBAL_ENV:
        GLOBAL_ENV[to_symbol(cn_name)] = GLOBAL_ENV[en_name]


# ============================================================
# 求值器
# ============================================================

def evaluate(exp, env: Env = GLOBAL_ENV):
    """求值表达式"""

    # None 直接返回
    if exp is None:
        return exp

    # 符号：从环境查找（但 :keyword 自求值，类 Common Lisp）
    if isinstance(exp, Symbol):
        if str(exp).startswith(':'):
            return exp
        try:
            scope = env.find(exp)
            if exp in scope:
                return scope[exp]
        except LookupError:
            scope = None
        # 兜底：中文别名查找链（如 用户输入数据 -> client-input 可能在执行时动态注入）
        target_cn = str(exp)
        if target_cn in _CN_ALIASES:
            target_sym = to_symbol(_CN_ALIASES[target_cn])
            if target_sym in GLOBAL_ENV:
                return GLOBAL_ENV[target_sym]
        raise NameError(f"未定义的符号: {exp}")

    # 非列表：常量直接返回
    if not isinstance(exp, list):
        return exp

    # 空列表返回 None
    if len(exp) == 0:
        return None

    # ========== 特殊形式 ==========

    op = exp[0]

    # check op 是否有效
    # - 如果 op 是宏，先展开
    if isinstance(op, Symbol):
        macro_name = op
        if macro_name in MACRO_TABLE:
            macro_def = MACRO_TABLE[macro_name]
            # 宏展开：用宏的参数模式匹配实参，替换模板
            expanded = _macro_expand(macro_def, exp[1:])
            # 展开后重新求值
            return evaluate(expanded, env)
        # 中文宏名也查一下
        cn_name = str(macro_name)
        if cn_name in _CN_MACRO_ALIASES:
            real_name = to_symbol(_CN_MACRO_ALIASES[cn_name])
            if real_name in MACRO_TABLE:
                macro_def = MACRO_TABLE[real_name]
                expanded = _macro_expand(macro_def, exp[1:])
                return evaluate(expanded, env)

    # 检查 op 是否有效（字符串或符号）
    # 注意：如果是列表（来自方括号解析的 lambda 表达式），会在下面处理

    # quote / 引
    if op == KW_QUOTE or op == KW_引:
        return exp[1]
    
    # if / 如果
    if op == KW_IF or op == KW_如果:
        _, test, conseq, *alt = exp
        alt = alt[0] if alt else None
        return evaluate(conseq if evaluate(test, env) else alt, env)

    # cond / 情况符合
    if op == KW_COND or op == KW_情况符合:
        for clause in exp[1:]:
            if len(clause) == 0:
                continue
            test = clause[0]
            # else 分支：如果条件是 'else', '否则', 或 '其它情况'
            if test == 'else' or test == '否则' or test == '其它情况':
                if len(clause) == 2:
                    return evaluate(clause[1], env)
                else:
                    return evaluate([KW_BEGIN] + clause[1:], env)
            # 普通分支
            if evaluate(test, env):
                if len(clause) == 2:
                    return evaluate(clause[1], env)
                else:
                    return evaluate([KW_BEGIN] + clause[1:], env)
        return None

    # define / 定义
    if op == KW_DEFINE or op == KW_定义:
        if isinstance(exp[1], list):
            # (define (f x) body) -> (define f (lambda (x) body))
            name, params = exp[1][0], exp[1][1:]
            body = exp[2] if len(exp) == 3 else [KW_BEGIN] + exp[2:]
            env[name] = Procedure(params, body, env)
        else:
            # (define name value) 或 (define name body1 body2 ...)
            _, name = exp[:2]
            if len(exp) > 3:
                # 多body形式，包装为begin
                value = evaluate([KW_BEGIN] + exp[2:], env)
            else:
                value = evaluate(exp[2], env)
            env[name] = value
        return None

    # defmacro / 定义宏
    if op == KW_DEFMACRO or op == KW_定义宏:
        _, name, params, *body = exp
        if isinstance(params, list):
            param_list = [to_symbol(str(p)) if isinstance(p, Symbol) else to_symbol(p) for p in params]
        elif isinstance(params, Symbol):
            param_list = [params]
        else:
            param_list = []
        template = body[0] if len(body) == 1 else [KW_BEGIN] + body
        MACRO_TABLE[name] = (param_list, template)
        _CN_MACRO_ALIASES[str(name)] = str(name)
        return None

    # macroexpand / 宏展开
    if op == KW_MACROEXPAND or op == KW_宏展开:
        _, name = exp
        if name in MACRO_TABLE:
            _, template = MACRO_TABLE[name]
            return template
        raise NameError(f"未定义的宏: {name}")

    # load / 引入
    if op == KW_LOAD or op == KW_引入:
        filepath = evaluate(exp[1], env)
        resolved = os.path.join(BASE_DIR, filepath) if not os.path.isabs(filepath) else filepath
        resolved = os.path.normpath(resolved)
        if resolved not in _loaded_files:
            _loaded_files.add(resolved)
            with open(resolved, 'r', encoding='utf-8') as f:
                code = f.read()
            run(code, env)
        return None

    # set! / 赋
    if op == KW_SET or op == KW_赋:
        _, var, val = exp
        env.find(var)[var] = evaluate(val, env)
        return None
    
    # lambda / 道 / 规定
    if op == KW_LAMBDA or op == KW_道 or op == KW_规定:
        _, params, *body = exp
        if isinstance(params, list):
            params = [to_symbol(str(p)) if isinstance(p, Symbol) else to_symbol(p) for p in params]
        elif isinstance(params, Symbol):
            params = [params]
        body = body[0] if len(body) == 1 else [KW_BEGIN] + body
        return Procedure(params, body, env)

    # begin / 开始
    if op == KW_BEGIN or op == KW_开始:
        result = None
        for e in exp[1:]:
            result = evaluate(e, env)
        return result
    
    # let / 令 / 命: (let ((x 1) (y 2)) body ...)
    if op == KW_LET or op == KW_令 or op == KW_命:
        _, bindings, *bodies = exp

        new_env = Env((), (), env)

        for b in bindings:
            if isinstance(b, list) and len(b) >= 2:
                name = b[0]
                value = evaluate(b[1], new_env)
                new_env[name] = value
            elif isinstance(b, list) and len(b) == 1:
                new_env[b[0]] = None

        # 执行所有 body 表达式，返回最后一个
        result = None
        for body in bodies:
            result = evaluate(body, new_env)
        return result
    
    # pipe: (pipe init fn1 fn2 ...)
    if op == KW_PIPE:
        _, init, *fns = exp
        result = evaluate(init, env)
        for fn_exp in fns:
            fn = evaluate(fn_exp, env)
            result = fn(result)
        return result
    
    # ->: (-> init (fn args...) (fn2 args...) ...)
    if op == KW_THREADED:
        _, init, *forms = exp
        result = evaluate(init, env)
        for form in forms:
            if isinstance(form, list):
                # 检查是否是 lambda 或其他特殊形式
                if form[0] == KW_LAMBDA or form[0] == KW_道:
                    # (lambda ...) 先求值得到函数，然后调用
                    fn = evaluate(form, env)
                    result = fn(result)
                else:
                    # (fn args...) -> (fn result args...)
                    fn = evaluate(form[0], env)
                    args = [evaluate(a, env) for a in form[1:]]
                    result = fn(result, *args)
            else:
                fn = evaluate(form, env)
                result = fn(result)
        return result

    # map / 映射 / 批处理: (map fn list)
    if op == KW_MAP or op == KW_映射 or op == KW_批处理:
        _, fn_exp, lst_exp = exp
        fn = evaluate(fn_exp, env)
        lst = evaluate(lst_exp, env)
        return [fn(x) for x in lst]

    # reduce / 归约 / 顺序执行: (reduce fn init list)
    if op == KW_REDUCE or op == KW_归约 or op == KW_顺序执行:
        _, fn_exp, init_exp, lst_exp = exp
        fn = evaluate(fn_exp, env)
        init = evaluate(init_exp, env)
        lst = evaluate(lst_exp, env)
        return functools.reduce(fn, lst, init)

    # filter / 过滤 / 筛选: (filter fn list)
    if op == KW_FILTER or op == KW_过滤 or op == KW_筛选:
        _, fn_exp, lst_exp = exp
        fn = evaluate(fn_exp, env)
        lst = evaluate(lst_exp, env)
        return [x for x in lst if fn(x)]

    # each / 遍历: (each fn list) - 对列表每项执行函数（副作用）
    if op == KW_EACH or op == KW_遍历:
        _, fn_exp, lst_exp = exp
        fn = evaluate(fn_exp, env)
        lst = evaluate(lst_exp, env)
        result = None
        for x in lst:
            result = fn(x)
        return result  # 返回最后一个结果

    # or / 或: 短路求值，遇到真值立即返回
    if op == KW_OR or op == KW_或:
        for arg in exp[1:]:
            val = evaluate(arg, env)
            if val:
                return val
        return False

    # and / 与: 短路求值，遇到假值立即返回
    if op == KW_AND or op == KW_与:
        val = True
        for arg in exp[1:]:
            val = evaluate(arg, env)
            if not val:
                return val
        return val

    # ========== 函数调用 ==========
    proc = evaluate(exp[0], env)
    if isinstance(proc, list) and len(proc) > 0:
        if isinstance(proc[0], Symbol) and proc[0] == KW_LAMBDA:
            proc = evaluate(proc, env)
    args = [evaluate(arg, env) for arg in exp[1:]]
    return proc(*args)


# ============================================================
# 辅助函数
# ============================================================

def to_lisp_str(exp) -> str:
    """将 Python 值转换为 Lisp 字符串表示"""
    if exp is True:
        return "#t"
    elif exp is False:
        return "#f"
    elif isinstance(exp, Symbol):
        return exp
    elif isinstance(exp, str):
        return f'"{exp}"'
    elif isinstance(exp, list):
        return '(' + ' '.join(map(to_lisp_str, exp)) + ')'
    else:
        return str(exp)


def run(code: str, env: Env = GLOBAL_ENV):
    """运行 Lisp 代码字符串"""
    from io import StringIO
    tokens = Tokenizer(StringIO(code))
    result = None
    while True:
        exp = parse(tokens)
        if exp is EOF:
            break
        result = evaluate(exp, env)
    return result


def run_file(path: str, env: Env = GLOBAL_ENV):
    """运行 Lisp 文件"""
    with open(path, 'r', encoding='utf-8') as f:
        code = f.read()
    return run(code, env)


def repl(prompt: str = "workflow-lisp> "):
    """交互式 REPL"""
    while True:
        try:
            code = input(prompt)
            if code.strip() == '':
                continue
            if code.strip() in ('exit', 'quit', '(exit)', '(quit)'):
                break
            result = run(code)
            if result is not None:
                print(to_lisp_str(result) if not isinstance(result, str) else result)
        except EOFError:
            break
        except Exception as e:
            print(f"错误：{format_user_error(e)}")


# ============================================================
# 主入口
# ============================================================

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        # 运行文件
        run_file(sys.argv[1])
    else:
        # 进入 REPL
        print("Lisp 工作流")
        print("输入 exit 退出")
        print()
        repl()
